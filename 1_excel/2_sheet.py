# 시트 관련 작업
from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet()  # 기본이름으로 새로운 시트 만들기. active는 현재 활성화된 시트를 가져오는것.

# 시트 이름 바꿔보기
ws.title = "My sheet"
ws.sheet_properties.tabColor = "ff66ff"  # 탭색상 바꾸기. rgb 형태로 값을 넣어준다.

ws1 = wb.create_sheet("your sheet")  # 생성하면서 타이틀 붙여주기.
ws2 = wb.create_sheet("new sheet", 2)  # index 넣어 생성 위치 지정.

new_ws = wb["new sheet"]  # dict 형태로 sheet에 접근이 가능.

print(wb.sheetnames)  # 모든 시트 이름 확인.

# sheet 복사
new_ws["A1"] = "Test"  # 문서 상의 a1셀의 값을 입력하는것.
target = wb.copy_worksheet(new_ws)  # 워크북 내의 new_ws를 복사.
target.title = "copied sheet"

wb.save("sample.xlsx")
wb.close()
