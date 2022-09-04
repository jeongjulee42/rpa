# 셀 범위 지정해서 가져오기
from openpyxl.utils.cell import coordinate_from_string
from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기.
ws.append(["번호", "영어", "수학"])  # 리스트 넣기.
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# 전체 rows
print(ws.rows)  # <generator object Worksheet._cells_by_row at 0x104220350> 이라는 오류 발생
print(tuple(ws.rows))  # 한줄씩 튜플로 묶어 가져온다.
print(tuple(ws.columns))  # 한열씩 튜플로 묶어 가져온다.

for row in tuple(ws.rows):
    print(row[1].value)

for column in tuple(ws.columns):
    print(column[0].value)

for row in ws.iter_rows():  # 전체 row에 대해 반복하며 가져오는것.
    print(row[1].value)

for column in ws.iter_cols():
    print(column[0].value)

# 이렇게 쓰면 범위 지정이 쉽다. iter_rows() 의 괄호에 값을 넣을수있다.
for row in ws.iter_rows(min_row=1,  max_row=5):  # 1~5까지 데이터 가져오기
    print(row[1].value)
# 1번째 줄부터 5번째 줄까지, 2번째 열부터 3번째 열까지.
for row in ws.iter_rows(min_row=1,  max_row=5, min_col=2, max_col=3):
    print(row[1].value)
wb.save("sample.xlsx")

# 한열씩 가지고 오고싶다 - iter_columns()
# 한줄씩 가지고 오고싶다 - iter_rows()
# min,max값은 지정 안하면 min-가장작은값, max-가장큰값.
