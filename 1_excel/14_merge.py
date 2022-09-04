# 셀 병합
import imp
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
# 병합
ws.merge_cells("B2:D2")  # b2~d2 까지 합친다
ws["B2"].value = "Merged Cell"


wb.save("sam_mer.xlsx")
