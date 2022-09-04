# 셀. 시트 내의 데이터를 편집하기.
from random import *
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "sht"

# 셀에 값 입력.
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3
ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

# 셀 값 읽어오기
print(ws["A1"])  # 셀 객체정보가 뜬다.
print(ws["A1"].value)  # 셀의 값을 출력
# 셀의 값이 없는 경우 none라고 나옴.

# 다른방법으로 셀에 접근 / row=1, column=1  == A1
print(ws.cell(row=1, column=1).value)

c = ws.cell(column=3, row=1, value=10)  # C1에 10을 넣는것.
print(c.value)

# 반복문을 이영한 랜덤 숫자 채우기.
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=randint(0, 100))  # 0~ 100사이의 숫자.


wb.save("sample.xlsx")
wb.close()

# ws.cell(row=? , column=? ).value 이렇게 써도 됨.
