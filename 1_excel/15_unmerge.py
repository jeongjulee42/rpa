# 병합된 셀 분할.
from openpyxl import load_workbook
wb = load_workbook("sam_mer.xlsx")
ws = wb.active

# B2~D2 병합셀 해제
ws.unmerge_cells("B2:D2")
wb.save("sam_unmer.xlsx")
