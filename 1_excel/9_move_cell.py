# 영어, 수학 위치 이동
from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 국어 영어 수학
ws.move_range("B1:C11", rows=0, cols=1)  # 범위 줄 열, -이면 왼쪽으로 이동한다.
ws["B1"].value = "국어"
ws.move_range("C1:D11", rows=5, cols=-1)

wb.save("sam_kor.xlsx")
