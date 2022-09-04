# 파일 열기
from openpyxl import load_workbook  # 파일 불러오기
wb = load_workbook("sample.xlsx")  # 파일에서 워크북 불러오기.
ws = wb.active  # 활성화된 시트 가져오기.

# cell 데이터 불러오기
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

# 줄 정보등을 모를때
for x in range(1, ws.max_row + 1):  # 최대 row 수
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()
