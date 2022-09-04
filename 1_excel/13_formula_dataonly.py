from openpyxl import load_workbook
# 수식 그대로 가져옴
# wb = load_workbook("sam_for.xlsx")
# ws = wb.active
# for row in ws.values:
#     for cell in row:
#         print(cell)


wb = load_workbook("sam_for.xlsx", data_only=True)  # 수식이 아닌 실제 데이터
# evaluate (계산) 되지 않은 상태의 데이터는 none이라고 표시/
# 파일 한번열었다가 닫을떄 저장하면 정상적으로 출력된다.
# 엑셀은 수식이 창을 마지막으로 열때 따로 계산된다.
ws = wb.active
for row in ws.values:
    for cell in row:
        print(cell)
