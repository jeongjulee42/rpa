
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

# 셀 너비 좁히기, a열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5
# 높이 50 설정
ws.row_dimensions[1].height = 50

# 스타일 적용 / from openpyxl.styles import Font
a1.font = Font(color="ff0000", italic=True, bold=True)
# name - 구체적인 폰트 명시, strike - 취소선 긋기
b1.font = Font(color="CC33FF", name="Arial", strike=True)
c1.font = Font(color="0000ff", size=20, underline="single")  # 글자크기, 밑줄 적용

# 테두리 설정 from openpyxl.styles import Border,Side
# 상하좌우 모두 얇은 테두리 적용 객체
thin_border = Border(left=Side(style="thin"), right=Side(
    style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 특정셀 배경색 바꾸기
# from openpyxl.styles import PatternFill
# 90점 넘는 셀에 대해 초록색으로 적용
# for row in ws.rows:
#     for cell in row:
#         if cell.column == 1:  # 번호열은 제외
#             continue

#         if isinstance(cell.value, int) and cell.value > 90:  # 제목을 제외
#             cell.fill = PatternFill(
#                 fgColor="00ff00", fill_type="solid")  # 배경 초록색으로
#             cell.font = Font(color="ff0000")  # 배경 색상도 변경

# 글자들 중앙 정렬
# from openpyxl.styles import Alignment
for row in ws.rows:
    for cell in row:
        # 각 셀에 대해 정렬.
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # center, left , right, top, bottom
        if cell.column == 1:
            continue

        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(
                fgColor="00ff00", fill_type="solid")
            cell.font = Font(color="ff0000")

# 틀 고정
ws.freeze_panes = "B2"  # b2 기준으로 틀 고정, 좌우로도 고정.

wb.save("sam_style.xlsx")
