# 셀 범위 지정해서 가져오기
from openpyxl.utils.cell import coordinate_from_string
from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기.
ws.append(["번호", "영어", "수학"])  # 리스트 넣기.
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"]  # 영어컬럼만 가지고 오기.
print(col_B)
for cell in col_B:
    print(cell.value)

col_range = ws["B:C"]  # b~c 까지의 모든 col 데이터 가지고 오기.
for cols in col_range:
    for cell in cols:
        print(cell.value)

row_title = ws[1]  # 1번째 row 만 가져오기
row_range = ws["2:7"]  # 7을 포함해서 가지고온다.
for rows in row_range:
    for row in rows:
        print(row.value)

# 어떤 데이터가 어떤 셀에있는지 정보를 필요로 할때
# from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2:ws.max_row]  # 2부터 마지막줄까지
for rows in row_range:
    for row in rows:
        # print(row.value, end=" ")
        # print(row.coordinate, end=" ")
        xy = coordinate_from_string(row.coordinate)  # 튜플형태로 나온다.
        print(xy, end=" ")
        print(xy[0])  # A
        print(xy[1])  # 1
    print()


wb.save("sample.xlsx")
