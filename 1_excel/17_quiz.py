from openpyxl import load_workbook
wb = load_workbook("score.xlsx")
ws = wb.active


def score(var):
    if var >= 90:
        return 'A'
    elif var >= 80:
        return 'B'
    elif var >= 70:
        return 'C'
    else:
        return 'D'


# 1. 퀴즈 2 점수 수정
for cell in ws["D"]:
    if isinstance(cell.value, int):
        cell.value = 10
# 2. 총점, 성적 정보 추가
ws["H1"] = "총점"
ws["I1"] = "성적"
for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=2, max_col=7)):
    tot = 0
    for cell in row:
        tot += cell.value
    num = "H" + str(idx+2)
    ws[num] = tot
    num = "I" + str(idx+2)
    ws[num] = score(tot)
    num1 = "B" + str(idx+2)
    if ws[num1].value < 5:
        ws[num] = 'F'
    # 3/ 출석 f

wb.save("scores.xlsx")
